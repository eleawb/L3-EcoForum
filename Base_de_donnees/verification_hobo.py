import psycopg2
import pandas as pd
import os
import csv
import sys 
import re
from pathlib import Path
import json
import argparse
from openpyxl import load_workbook
from datetime import datetime
from dotenv import load_dotenv #ajout

#pour que sur windows ou mac les chemins fichiers soient les mêmes
def normaliser_chemin(chemin):
    """Convertit un chemin Windows en chemin valide pour l'OS courant"""
    if not chemin:
        return chemin
    
    #remplacer les \ par des /
    chemin = chemin.replace('\\', '/')
    
    #enlever le './' ou '.\\' au début si présent
    if chemin.startswith('./') or chemin.startswith('.\\'):
        chemin = chemin[2:]
    elif chemin.startswith('.'):
        chemin = chemin[1:]
    
    #sur Mac/Linux, depuis Base_de_donnees, il faut remonter d'un dossier
    if sys.platform != "win32":
        #si le chemin ne commence pas déjà par ../
        if not chemin.startswith('/') and not chemin.startswith('../'):
            chemin = './' + chemin
    
    return chemin


def expand_item(item):
    """
    Fonction écrite par une IA car je n'arrivais pas à m'en sortir seule et que c'est une fonction utile, mais secondaire pour la vérification des fichiers
    Principe : passer de ["#", "date et heure (CET/CEST) || CET || CEST", "température , °C"] à ["#", ["date et heure (CET/CEST)", "date et heure CET", "date et heure CEST"], "température , °C"]
    La fonction prend en entrée une chaîne de caractères et teste si elle contient "||"
    Si elle contient "||", cela veut dire qu'il y a plusieurs noms possibles pour une même colonne dans plusieurs fichiers et qu'il faut donc tester toutes les possibilités avant de rejeter le fichier
    Si elle contient "||", il faut tout d'abord séparer la chaîne avec comme séparateur "||"
    Ensuite, il faut cherche la base, c'est-à-dire la string située avant la permière string qui était suivie par "||".
    Dans notre exemple, la base est "date et heure"
    Les variantes sont toutes les strings juste avant un "||" et/ou juste après un "||".
    La première string suivie par un "||" est la première variante.
    Il est possible que la dernière variante soit suivie d'une string que nous appelons la fin.
    Dans notre exemple, la fin est ""
    Une fois que la base et la fin sont séparées des variantes, il faut construire la liste de toutes les combinaisons de la forme "base variante fin"
    """
    if "||" not in item:
        return item

    # Étape 1 : découper en morceaux bruts
    parts = item.split("||")

    base = parts[0].strip()                     # avant le premier ||
    raw_variants = [p.strip() for p in parts[1:]]  # tout le reste

    # Étape 2 : la dernière partie contient la dernière variante + éventuellement la fin
    last = raw_variants[-1]

    # On sépare la dernière variante et la fin :
    # On cherche la première occurrence de la dernière variante dans 'last'
    # puis tout ce qui suit est la fin.
    last_variant = last.split()[0]  # première "mot" = variante brute
    idx = last.find(last_variant)

    # Variante finale
    final_variant = last_variant

    # Fin = tout ce qui suit la variante finale
    fin = last[idx + len(last_variant):].strip()

    # Variantes = toutes sauf la dernière, plus la variante finale propre
    variants = raw_variants[:-1] + [final_variant]

    # Étape 3 : zone à remplacer = dernier bloc de la base
    zone = base.split()[-1]

    # Étape 4 : construire les résultats
    results = []

    # Version originale
    results.append(base + " " + fin if fin else base)

    # Versions remplacées
    for v in variants:
        new = base.replace(zone, v)
        results.append(new + (" " + fin if fin else ""))

    return results


def verification_hobo(metajson):
    """
    Prend en entrée un fichier json de la forme suivante :
    {"script": "verification", 
    "fichier_mesure": "chemin\\(n\u00b036) 21149857 2026-01-28 15_37_05 CET.xlsx",
    "nom_outil": "HOBO", 
    "num_instrument": "HOBO_n°36"}
    """

    with open(metajson, "r", encoding="utf-8") as f:
        data = json.load(f)

        #normalisation du chemin
        data["chemin_source"] = normaliser_chemin(data["chemin_source"])

    dico = {
        "numero_serie" : "",
        "extension" : "",
        "date_recueil" : "",
        "reussite" : False,
        "commentaire" : "",
        "date_import" : "",
        "type_source" : "",
    }

    #pour obtenir le nom du fichier avec son extension sans avoir le chemin avant
    nomFic = Path(data["chemin_source"]).name
    #print(nomFic)
    num_instru = data["num_instrument"]
    #print(num_instru)
    nom_instru = data["nom_outil"]
    #print(nom_instru)

    #json.dump(dico, f, indent=4, ensure_ascii=False)

    cur.execute("SELECT * FROM structure_fichier sf JOIN instrument_mesure im ON lower(im.nom_outil) = lower(sf.nom_instrument) WHERE lower(im.num_instrument) = lower(%s);", (num_instru,))
    rows = cur.fetchall()
    #print(rows)
    for row in rows :
        if re.search(row[5], nomFic):
            #print("bon nom de fichier")
            #colonnes = pd.read_excel(data["fichier_mesure"], nrows=0).columns #avec pandas, mais charge tout le fichier, ce qui n'est pas utile et pas otpimal
            wb = load_workbook(data["chemin_source"], read_only=True)
            ws = wb[wb.sheetnames[0]] #accès au premier onglet du fichier
            colonnes = ws.max_column
            if row[3] == colonnes:
                #print("bon nombre de colonnes")
                entetes = next(ws.iter_rows(max_row=1, values_only=True))#lis la ligne des en-têtes (donne un tuple)
                entetes = list(entetes)#on transforme le tuple en liste ?

                cols = row[2].split("; ")
                lcols = []
                for c in cols:
                    lcols.append(expand_item(c))
                #print(lcols)
                
                for i in range(colonnes):
                    if isinstance(lcols[i], str):
                        if lcols[i].lower().strip() != entetes[i].lower().strip():
                            #print("nom de colonne différent str")
                            dico["commentaire"] = "nom de la colonne n°" + i + "différent, nom attendu : " + lcols[i]
                            #return dico
                            with open("retour.json", "w", encoding="utf-8") as f:
                                json.dump(dico, f, indent=4, ensure_ascii=False)
                            print(os.path.join(os.getcwd(), "retour.json"))
                        #else:
                            #print("bon nom de colonne str")
                    elif isinstance(lcols[i], list):
                        memecol = False
                        for j in range(len(lcols[i])):
                            #print(lcols[i][j].lower().strip())
                            #print(entetes[i].lower().strip())
                            if lcols[i][j].lower().strip() == entetes[i].lower().strip():
                                memecol = True
                                #print("bon nom de colonne list")
                        if memecol == False:
                            #print("nom de colonne différent list")
                            dico["commentaire"] = "nom de la colonne n°" + i + "différent, noms attendus : " + lcols[i]
                            with open("retour.json", "w", encoding="utf-8") as f:
                                json.dump(dico, f, indent=4, ensure_ascii=False)
                            print(os.path.join(os.getcwd(), "retour.json"))
            else :
                dico["commentaire"] = "le fichier ne contient pas le bon nombre de colonnes, nombre attendu : " + row[3]
                with open("retour.json", "w", encoding="utf-8") as f:
                    json.dump(dico, f, indent=4, ensure_ascii=False)
                print(os.path.join(os.getcwd(), "retour.json"))
        else:
            dico["commentaire"] = "Le nom du fichier n'a pas le bon format, format attendu : " + row[5]
            with open("retour.json", "w", encoding="utf-8") as f:
                json.dump(dico, f, indent=4, ensure_ascii=False)
            print(os.path.join(os.getcwd(), "retour.json"))

    dico["reussite"] = True
    dico["extension"] = rows[0][1]
    dico["numero_serie"] = rows[0][11]
    dico["date_import"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pattern = r"([0-9]{4}-[0-9]{2}-[0-9]{2} [0-9]{2}_[0-9]{2}_[0-9]{2}) (?=CET\.xlsx)"

    match = re.search(pattern, data["chemin_source"])

    if match:
        raw_dt = match.group(1)
    #print("Brut :", raw_dt)

    # Conversion des underscores en deux-points
    clean_dt = raw_dt.replace("_", ":")
    #print("Format standard :", clean_dt)
    dico["date_recueil"] = clean_dt
    dico["type_source"] = "fichier_mesure"
    with open("retour.json", "w", encoding="utf-8") as f:
        json.dump(dico, f, indent=4, ensure_ascii=False)
    print(os.path.join(os.getcwd(), "retour.json"))


if __name__ == "__main__":

    parser = argparse.ArgumentParser()
    parser.add_argument("--json")
    args = parser.parse_args()

    if len(sys.argv) != 2:
        dico = {}
        dico["commentaire"] = "Usage : ./verification_hobo.py JSON/verification_Hobo.json"
        dico["reussite"] = False
        with open("retour.json", "w", encoding="utf-8") as f:
            json.dump(dico, f, indent=4, ensure_ascii=False)
        print(os.path.join(os.getcwd(), "retour.json"))

    else :


        load_dotenv()


    # Connexion à la base
    #conn = psycopg2.connect(
        #host="localhost",
        #database="eco_forum",
        #user="postgres",
        #password="123456",
        #port=5432
    #)
        conn = psycopg2.connect(
            host=os.getenv("DB_HOST", "localhost"),
            database=os.getenv("DB_NAME", "eco_forum"),
            user=os.getenv("DB_USER", "postgres"),
            password=os.getenv("DB_PASSWORD", ""),
            port=os.getenv("DB_PORT", 5432)
        )

        #Création du curseur qui nous permettra de faire les requêtes
        cur = conn.cursor()

        verification_hobo(args.json)

        # Fermeture curseur et connexion à la base
        cur.close()
        conn.close()
